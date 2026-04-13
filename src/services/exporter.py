import os
import logging
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Dict, Union, List

import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Módulos del proyecto
from src.models.plant import Plant
from src.core.database import DatabaseManager

class DataExporter:
    def __init__(self, db_manager: DatabaseManager, plant: Plant):
        self.db = db_manager
        self.plant = plant
        self.out_dir = self.plant.output_path
        self.img_dir = self.plant.img_path
        self.today_str = date.today().strftime('%Y-%m-%d')
        
        # Crear carpetas de salida
        os.makedirs(self.out_dir, exist_ok=True)
        os.makedirs(self.img_dir, exist_ok=True)

    # =========================================================================
    # --- MOTOR DE EXPORTACIÓN ---
    # =========================================================================
    def _export_excel(self, data: Union[pd.DataFrame, Dict[str, pd.DataFrame]], filename: str, is_error_report=False):
        """Genera archivos Excel con formato avanzado y soporte multi-hoja."""
        dict_dfs = data if isinstance(data, dict) else {'Datos': data}
        dict_dfs = {k: v for k, v in dict_dfs.items() if v is not None and not v.empty}
        
        if not dict_dfs:
            logging.warning(f"      ! No hay datos para generar: {filename}")
            return
        
        full_path = os.path.join(self.out_dir, filename)
        
        try:
            with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
                for sheet_name, df_sheet in dict_dfs.items():
                    df_copy = df_sheet.copy()
                    
                    # Limpieza de Timezones para Excel
                    for col in df_copy.select_dtypes(include=['datetimetz']).columns:
                        df_copy[col] = df_copy[col].dt.tz_localize(None)
                    
                    # Formateo de decimales a estilo español (coma)
                    for col in df_copy.columns:
                        if any(isinstance(x, Decimal) for x in df_copy[col].dropna()):
                            df_copy[col] = df_copy[col].apply(lambda x: str(x).replace('.', ',') if isinstance(x, Decimal) else x)

                    sheet_safe = sheet_name[:31]
                    df_copy.to_excel(writer, index=False, sheet_name=sheet_safe)
                    worksheet = writer.sheets[sheet_safe]
                    
                    # Estilos visuales
                    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    
                    # Encabezados y ajuste de columnas
                    for col_idx, col_name in enumerate(df_copy.columns, start=1):
                        cell = worksheet.cell(row=1, column=col_idx)
                        cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, center_align, thin_border
                        max_len = max((df_copy[col_name].astype(str).map(len).max() if not df_copy.empty else 0), len(str(col_name)))
                        worksheet.column_dimensions[get_column_letter(col_idx)].width = max(min(max_len + 2, 50), 12)

                    # Cuerpo de la tabla
                    for row_idx in range(2, len(df_copy) + 2):
                        for col_idx in range(1, len(df_copy.columns) + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.alignment, cell.border = center_align, thin_border

                    worksheet.auto_filter.ref = worksheet.dimensions 
                    worksheet.freeze_panes = 'A2'

                    # Formato condicional para reporte de errores
                    if is_error_report and 'tipo_de_error' in df_copy.columns:
                        col_tipo_idx = df_copy.columns.get_loc('tipo_de_error') + 1
                        for row_idx in range(2, len(df_copy) + 2):
                            cell = worksheet.cell(row=row_idx, column=col_tipo_idx)
                            if cell.value == 'Error':
                                cell.fill, cell.font = PatternFill(start_color="FFC7CE", fill_type="solid"), Font(color="9C0006", bold=True)
                            elif cell.value == 'Alarma':
                                cell.fill, cell.font = PatternFill(start_color="FFEB9C", fill_type="solid"), Font(color="9C6500", bold=True)

            logging.info(f"Archivo generado: {filename}")
        except Exception as e:
            logging.error(f"Fallo al guardar {filename}: {e}")

    # =========================================================================
    # --- REPORTES ---
    # =========================================================================

    def export_excel_reports(self):
        """Orquestador de todos los reportes Excel."""
        logging.info(f"[{self.plant.name}] Generando reportes Excel...")
        
        # 1. Reporte de Datos Validados
        query_val = "SELECT * FROM public.validated_data WHERE DATE(created_at) = CURRENT_DATE ORDER BY timestamp;"
        df_val = pd.read_sql(query_val, self.db.conn)
        if not df_val.empty:
            df_val.drop(columns=['id', 'raw_data_id', 'status', 'created_at', 'processed_at'], inplace=True, errors='ignore')
            self._export_excel(df_val, f"reporte_validados_{self.today_str}.xlsx")

        # 2. Reporte de Errores
        query_err = """
            SELECT d.timestamp, e.error_type as tipo_de_error, COALESCE(vr.error_message, 'Error') AS mensajes_de_error,
                   e.offending_column as columnas_con_error, e.offending_value as valores_con_error
            FROM public.validation_error_by_rules e
            JOIN public.raw_data d ON e.raw_data_id = d.id
            LEFT JOIN public.validation_rules vr ON e.validation_rule_id = vr.id
            WHERE DATE(e.created_at) = CURRENT_DATE ORDER BY d.timestamp;
        """
        df_err = pd.read_sql(query_err, self.db.conn)
        if not df_err.empty:
            df_err['tipo_de_error'] = df_err['tipo_de_error'].replace({'error': 'Error', 'alarm': 'Alarma'})
            self._export_excel(df_err, f"reporte_errores_{self.today_str}.xlsx", is_error_report=True)

        # 3. Reporte de Exclusiones (Múltiples hojas)
        self.export_exclusions_report()

        # 4. Reporte de Días Faltantes (Solo fechas)
        self.export_missing_days_report()

    def export_exclusions_report(self):
        """Crea el reporte de exclusiones segmentado por tipo."""
        query_exc = """
            SELECT form_timestamp as fecha_formulario, exclusion_start as inicio_exclusion, 
                   exclusion_end as fin_exclusion, exclusion_type as tipo_exclusion, 
                   motivo, observacion, potencia_pico_kw, excluded_variables as variables_excluidas
            FROM public.excluded_data ORDER BY form_timestamp DESC;
        """
        df_exc = pd.read_sql(query_exc, self.db.conn)
        if df_exc.empty: return

        hojas = {}
        mapping = {
            'Periodo_Marcado': 'Exclusión de periodo marcado',
            'Variable_Fuera_Rango': 'Aceptación de una variable fuera de rango',
            'Modificacion_Potencia': 'Modificación de la potencia pico'
        }
        
        for sheet, type_str in mapping.items():
            df_sub = df_exc[df_exc['tipo_exclusion'] == type_str].copy()
            if not df_sub.empty:
                if sheet == 'Periodo_Marcado': df_sub = df_sub.drop(columns=['potencia_pico_kw', 'variables_excluidas'])
                elif sheet == 'Variable_Fuera_Rango': df_sub = df_sub.drop(columns=['potencia_pico_kw'])
                elif sheet == 'Modificacion_Potencia': df_sub = df_sub.drop(columns=['variables_excluidas'])
                hojas[sheet] = df_sub

        self._export_excel(hojas, f"reporte_exclusiones_{self.today_str}.xlsx")

    def export_missing_days_report(self):
        """Identifica fechas sin carga en load_control y exporta solo la columna de fecha."""
        logging.info(f"[{self.plant.name}] Buscando fechas faltantes...")
        query = "SELECT inventory_date FROM public.load_control ORDER BY inventory_date ASC;"
        df_inv = pd.read_sql(query, self.db.conn)
        
        if df_inv.empty: return

        fechas_cargadas = set(df_inv['inventory_date'])
        fecha_inicio = min(fechas_cargadas)
        fecha_fin = date.today()

        faltantes = []
        curr = fecha_inicio
        while curr <= fecha_fin:
            if curr not in fechas_cargadas:
                faltantes.append(curr)
            curr += timedelta(days=1)
        
        if faltantes:
            df_miss = pd.DataFrame(faltantes, columns=['fecha_no_cargada'])
            self._export_excel(df_miss, f"reporte_dias_faltantes_{self.today_str}.xlsx")
        else:
            logging.info("No hay días faltantes registrados.")

    def export_graphics(self):
        """Genera gráficos de barras para errores detectados hoy."""
        query = """
            SELECT ve.offending_column AS nombre_columna, COUNT(*) AS cantidad, ve.error_type
            FROM validation_error_by_rules ve
            WHERE DATE(ve.created_at) = CURRENT_DATE
            GROUP BY ve.offending_column, ve.error_type
        """
        df = pd.read_sql(query, self.db.conn)
        if df.empty: return

        plt.style.use('ggplot')
        for err_type, title, color in [('error', 'Errores', '#c0392b'), ('alarm', 'Alarmas', '#f39c12')]:
            df_sub = df[df['error_type'] == err_type].copy()
            if df_sub.empty: continue 
            
            df_sub = df_sub.sort_values(by='cantidad', ascending=True)
            plt.figure(figsize=(10, max(6, len(df_sub) * 0.4)))
            plt.barh(df_sub['nombre_columna'], df_sub['cantidad'], color=color)
            plt.title(f"{title} por Columna - {self.today_str}")
            plt.tight_layout()
            plt.savefig(os.path.join(self.img_dir, f"grafico_{title.lower()}_bar.png"))
            plt.close()

    def run(self):
        """Ejecución principal del exportador."""
        try:
            self.export_excel_reports()
            self.export_graphics()
        except Exception as e:
            logging.error(f"Error general en ejecución: {e}")